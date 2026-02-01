"""
재고관리 자동화 스크립트
조흥지에프(공급처코드: 20001787) 재고 데이터 자동 업데이트

작성일: 2025-01-23
작성자: AI Assistant for Lucy
"""

import openpyxl
import pandas as pd
from datetime import datetime
import re
import os
import sys


class InventoryUpdater:
    def __init__(self, file_path):
        """
        초기화
        :param file_path: 재고관리.xlsx 파일 경로
        """
        self.file_path = file_path
        self.wb = None
        self.supplier_code = 20001787  # 조흥지에프
        self.target_sheet_name = '조흥'
        self.log_messages = []
        
    def log(self, message):
        """로그 메시지 기록"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_msg = f"[{timestamp}] {message}"
        print(log_msg)
        self.log_messages.append(log_msg)
    
    def find_latest_raw_sheet(self):
        """
        최신 날짜의 raw data 시트 찾기
        :return: 시트 이름 (예: '0122')
        """
        self.log("최신 raw data 시트 검색 중...")
        
        # 날짜 형식의 시트 찾기 (MMDD 형식)
        date_sheets = []
        for sheet_name in self.wb.sheetnames:
            # 4자리 숫자로 된 시트명 찾기
            if re.match(r'^\d{4}$', sheet_name):
                try:
                    # MMDD 형식을 datetime으로 변환
                    month = int(sheet_name[:2])
                    day = int(sheet_name[2:])
                    year = datetime.now().year
                    date = datetime(year, month, day)
                    date_sheets.append((sheet_name, date))
                except ValueError:
                    continue
        
        if not date_sheets:
            raise ValueError("날짜 형식의 시트를 찾을 수 없습니다 (MMDD 형식)")
        
        # 가장 최근 날짜 시트 선택
        latest_sheet = max(date_sheets, key=lambda x: x[1])
        self.log(f"최신 시트 발견: {latest_sheet[0]} ({latest_sheet[1].strftime('%m월 %d일')})")
        
        return latest_sheet[0]
    
    def extract_joheung_data(self, sheet_name):
        """
        raw data 시트에서 조흥지에프 데이터 추출
        :param sheet_name: raw data 시트 이름
        :return: DataFrame
        """
        self.log(f"'{sheet_name}' 시트에서 조흥지에프 데이터 추출 중...")
        
        # 시트 읽기
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        
        # 공급처코드로 필터링
        if '공급처코드' not in df.columns:
            raise ValueError(f"'{sheet_name}' 시트에 '공급처코드' 컬럼이 없습니다.")
        
        joheung_df = df[df['공급처코드'] == self.supplier_code].copy()
        
        if joheung_df.empty:
            self.log(f"⚠️ 경고: '{sheet_name}' 시트에서 공급처코드 {self.supplier_code} 데이터를 찾을 수 없습니다.")
            return None
        
        self.log(f"✓ {len(joheung_df)}개 품목 데이터 추출 완료")
        
        # 필요한 컬럼만 추출
        columns_to_extract = {
            '품목코드': '품목코드',
            '당월\n 판매량': '당월_판매량',
            '전월\n 판매량': '전월_판매량',
            '전전월\n 판매량': '전전월_판매량',
            '합계': '합계_재고',
            '출고가능량\n (가입고포함)': '출고가능량'
        }
        
        extracted_data = {}
        for orig_col, new_col in columns_to_extract.items():
            if orig_col in joheung_df.columns:
                extracted_data[new_col] = joheung_df[orig_col]
            else:
                self.log(f"⚠️ 경고: '{orig_col}' 컬럼을 찾을 수 없습니다.")
        
        result_df = pd.DataFrame(extracted_data)
        return result_df
    
    def update_joheung_sheet(self, raw_data, update_date):
        """
        조흥 시트 업데이트
        :param raw_data: raw data에서 추출한 DataFrame
        :param update_date: 업데이트 날짜 (예: '0122')
        """
        if raw_data is None or raw_data.empty:
            self.log("업데이트할 데이터가 없습니다.")
            return
        
        self.log(f"'{self.target_sheet_name}' 시트 업데이트 중...")
        
        # 조흥 시트 읽기 (header=1: 2번째 행이 헤더)
        df_joheung = pd.read_excel(self.file_path, sheet_name=self.target_sheet_name, header=1)
        
        # openpyxl로 직접 접근하여 업데이트
        ws = self.wb[self.target_sheet_name]
        
        # 품목코드 컬럼 위치 찾기 (B열: 2번째 컬럼)
        품목코드_col = 2  # B열
        
        # 업데이트할 컬럼 위치 매핑 (조흥 시트 기준)
        column_mapping = {
            '재고1/21': 7,     # G열 (날짜에 따라 변경 가능)
            '당월\n 판매량': 8,   # H열
            '전월\n 판매량': 9,   # I열
            '전전월\n 판매량': 10, # J열
            '현재고': 19        # S열
        }
        
        updated_count = 0
        new_items = []
        
        # raw_data의 각 품목에 대해
        for idx, row in raw_data.iterrows():
            품목코드 = row['품목코드']
            
            # 조흥 시트에서 해당 품목코드 찾기
            found = False
            for ws_row in range(3, ws.max_row + 1):  # 3행부터 (1행: 공급처정보, 2행: 헤더)
                if ws.cell(row=ws_row, column=품목코드_col).value == 품목코드:
                    found = True
                    
                    # 데이터 업데이트
                    # 당월 판매량
                    if '당월_판매량' in row and pd.notna(row['당월_판매량']):
                        ws.cell(row=ws_row, column=column_mapping['당월\n 판매량']).value = row['당월_판매량']
                    
                    # 전월 판매량
                    if '전월_판매량' in row and pd.notna(row['전월_판매량']):
                        ws.cell(row=ws_row, column=column_mapping['전월\n 판매량']).value = row['전월_판매량']
                    
                    # 전전월 판매량
                    if '전전월_판매량' in row and pd.notna(row['전전월_판매량']):
                        ws.cell(row=ws_row, column=column_mapping['전전월\n 판매량']).value = row['전전월_판매량']
                    
                    # 재고 (합계)
                    if '합계_재고' in row and pd.notna(row['합계_재고']):
                        ws.cell(row=ws_row, column=column_mapping['재고1/21']).value = row['합계_재고']
                    
                    # 현재고 (출고가능량)
                    if '출고가능량' in row and pd.notna(row['출고가능량']):
                        ws.cell(row=ws_row, column=column_mapping['현재고']).value = row['출고가능량']
                    
                    updated_count += 1
                    break
            
            if not found:
                new_items.append(품목코드)
        
        self.log(f"✓ {updated_count}개 품목 업데이트 완료")
        
        if new_items:
            self.log(f"⚠️ 조흥 시트에 없는 신규 품목 발견: {len(new_items)}개")
            for item in new_items[:5]:  # 최대 5개만 출력
                self.log(f"   - {item}")
            if len(new_items) > 5:
                self.log(f"   ... 외 {len(new_items) - 5}개")
    
    def save_file(self, backup=True):
        """
        파일 저장
        :param backup: 백업 생성 여부
        """
        if backup:
            # 백업 파일 생성
            backup_path = self.file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
            self.log(f"백업 파일 생성: {backup_path}")
            self.wb.save(backup_path)
        
        # 원본 파일 저장
        self.log("파일 저장 중...")
        self.wb.save(self.file_path)
        self.log("✓ 파일 저장 완료")
    
    def save_log(self):
        """로그 파일 저장"""
        log_dir = os.path.dirname(self.file_path) or '.'
        log_file = os.path.join(log_dir, f'update_log_{datetime.now().strftime("%Y%m%d")}.txt')
        
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write('\n' + '='*80 + '\n')
            for msg in self.log_messages:
                f.write(msg + '\n')
        
        self.log(f"로그 파일 저장: {log_file}")
    
    def run(self):
        """메인 실행 함수"""
        try:
            self.log("="*80)
            self.log("재고관리 자동화 스크립트 시작")
            self.log("="*80)
            
            # 1. 파일 열기
            self.log(f"파일 열기: {self.file_path}")
            self.wb = openpyxl.load_workbook(self.file_path)
            
            # 2. 최신 raw data 시트 찾기
            latest_sheet = self.find_latest_raw_sheet()
            
            # 3. 조흥지에프 데이터 추출
            raw_data = self.extract_joheung_data(latest_sheet)
            
            # 4. 조흥 시트 업데이트
            self.update_joheung_sheet(raw_data, latest_sheet)
            
            # 5. 파일 저장
            self.save_file(backup=True)
            
            # 6. 로그 저장
            self.save_log()
            
            self.log("="*80)
            self.log("✓ 모든 작업 완료!")
            self.log("="*80)
            
        except Exception as e:
            self.log(f"❌ 오류 발생: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            self.save_log()
            raise
        
        finally:
            if self.wb:
                self.wb.close()


def main():
    """메인 함수"""
    # 파일 경로 설정
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # 기본 경로: 스크립트와 같은 폴더의 재고관리.xlsx
        file_path = os.path.join(os.path.dirname(__file__), '재고관리.xlsx')
    
    # 파일 존재 확인
    if not os.path.exists(file_path):
        print(f"❌ 오류: 파일을 찾을 수 없습니다 - {file_path}")
        print("\n사용법:")
        print("  python update_joheung_inventory.py [파일경로]")
        print("\n예시:")
        print("  python update_joheung_inventory.py C:\\Users\\Lucy\\재고관리.xlsx")
        sys.exit(1)
    
    # 업데이터 실행
    updater = InventoryUpdater(file_path)
    updater.run()
    
    # 완료 후 대기 (콘솔창 바로 닫히지 않도록)
    input("\n아무 키나 누르면 종료됩니다...")


if __name__ == "__main__":
    main()
